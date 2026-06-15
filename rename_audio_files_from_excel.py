from __future__ import annotations

import argparse
import re
from pathlib import Path


INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1F]')
WHITESPACE_RE = re.compile(r"\s+")
RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    "COM1",
    "COM2",
    "COM3",
    "COM4",
    "COM5",
    "COM6",
    "COM7",
    "COM8",
    "COM9",
    "LPT1",
    "LPT2",
    "LPT3",
    "LPT4",
    "LPT5",
    "LPT6",
    "LPT7",
    "LPT8",
    "LPT9",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Rename WAV files based on rows from wav_text_normalized_3_columns.xlsx "
            "and write a new Excel file with the updated names."
        )
    )
    parser.add_argument(
        "--input",
        default="wav_text_normalized_3_columns.xlsx",
        help="Input Excel file. Default: wav_text_normalized_3_columns.xlsx",
    )
    parser.add_argument(
        "--output",
        default="wav_text_renamed.xlsx",
        help="Output Excel file. Default: wav_text_renamed.xlsx",
    )
    parser.add_argument(
        "--wav-dir",
        default="wavs",
        help="Folder containing the WAV files. Default: wavs",
    )
    parser.add_argument(
        "--sheet",
        help="Input sheet name. If omitted, the active sheet is used.",
    )
    parser.add_argument(
        "--name-source",
        default="sequential",
        choices=("sequential", "normalize text", "old text", "audio file", "row"),
        help=(
            "Which value to use when generating the new filename. "
            "Default: sequential"
        ),
    )
    parser.add_argument(
        "--base-name",
        default="racordsFinal",
        help="Base name for sequential filenames. Default: racordsFinal",
    )
    return parser.parse_args()


def require_dependencies():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        missing = exc.name or "openpyxl"
        raise SystemExit(
            f"Missing dependency: {missing}\n"
            "Install dependencies with:\n"
            "  pip install openpyxl"
        ) from exc

    return Workbook, load_workbook


def resolve_path(base_dir: Path, value: str) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return base_dir / path


def header_map(sheet) -> dict[str, int]:
    headers = {}
    for column_number in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=column_number).value
        if value is not None:
            headers[str(value).strip()] = column_number
    return headers


def sanitize_filename(value: str) -> str:
    text = INVALID_FILENAME_CHARS.sub("", str(value)).strip()
    text = WHITESPACE_RE.sub(" ", text).replace(" ", "_")
    text = re.sub(r"_+", "_", text)
    text = text.strip("._ ")
    return text


def is_reserved_name(stem: str) -> bool:
    return stem.upper() in RESERVED_NAMES


def unique_target_path(wav_dir: Path, desired_name: str, used_names: set[str]) -> Path:
    candidate = desired_name
    stem = Path(candidate).stem
    suffix = Path(candidate).suffix
    counter = 1

    while True:
        target = wav_dir / candidate
        if candidate.lower() not in used_names and not target.exists():
            used_names.add(candidate.lower())
            return target

        counter += 1
        candidate = f"{stem}_{counter}{suffix}"


def extract_source_value(row_number: int, sheet, headers: dict[str, int], name_source: str) -> str:
    if name_source == "row":
        return f"row_{row_number}"

    column = headers.get(name_source)
    if not column:
        return f"row_{row_number}"

    value = sheet.cell(row=row_number, column=column).value
    return str(value or "").strip()


def build_new_filename(
    row_number: int,
    sheet,
    headers: dict[str, int],
    source_name: str,
    base_name: str,
    sequence_number: int,
) -> str:
    if source_name == "sequential":
        candidate = sanitize_filename(base_name) or "record"
        if is_reserved_name(candidate):
            candidate = f"{candidate}_file"
        candidate = f"{candidate}_{sequence_number}"
    else:
        candidate = extract_source_value(row_number, sheet, headers, source_name)
        if source_name != "audio file" and (
            not candidate or candidate.startswith("ERROR:") or candidate.startswith("SKIPPED:")
        ):
            candidate = extract_source_value(row_number, sheet, headers, "audio file")

        if source_name == "audio file":
            candidate = Path(candidate).stem

        candidate = sanitize_filename(candidate)
        if not candidate:
            candidate = f"row_{row_number}"

        if is_reserved_name(candidate):
            candidate = f"{candidate}_file"

    if not Path(candidate).suffix.lower() == ".wav":
        candidate = f"{candidate}.wav"

    if len(candidate) > 120:
        stem = Path(candidate).stem[:110]
        candidate = f"{stem}.wav"

    return candidate


def print_progress(done: int, total: int, current: str = "") -> None:
    width = 30
    filled = int(width * done / total) if total else width
    bar = "#" * filled + "-" * (width - filled)
    percent = int(100 * done / total) if total else 100
    message = f"\rProgress: [{bar}] {done}/{total} ({percent}%)"
    if current:
        message += f" {current}"
    print(message[:120].ljust(120), end="", flush=True)


def main() -> None:
    args = parse_args()
    Workbook, load_workbook = require_dependencies()

    base_dir = Path(__file__).resolve().parent
    input_path = resolve_path(base_dir, args.input)
    output_path = resolve_path(base_dir, args.output)
    wav_dir = resolve_path(base_dir, args.wav_dir)

    if not input_path.exists():
        raise SystemExit(f"Input Excel file not found: {input_path}")
    if not wav_dir.exists():
        raise SystemExit(f"WAV folder not found: {wav_dir}")

    workbook = load_workbook(input_path)
    sheet = workbook[args.sheet] if args.sheet else workbook.active

    headers = header_map(sheet)
    required_headers = {"audio file", "old text", "normalize text"}
    missing_headers = sorted(required_headers - set(headers))
    if missing_headers:
        raise SystemExit(
            f"Missing required column(s) in {input_path}: {', '.join(missing_headers)}"
        )

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "renamed_audio_files"
    output_sheet.append(["audio file", "old text", "normalize text"])

    used_names = {path.name.lower() for path in wav_dir.iterdir() if path.is_file()}
    renamed_count = 0
    skipped_missing = 0
    sequence_number = 1

    total_rows = max(0, sheet.max_row - 1)
    print_progress(0, total_rows)

    for index, row_number in enumerate(range(2, sheet.max_row + 1), start=1):
        original_name = str(sheet.cell(row=row_number, column=headers["audio file"]).value or "").strip()
        if not original_name:
            skipped_missing += 1
            print_progress(index, total_rows, "missing audio file name")
            continue

        source_path = wav_dir / original_name
        if not source_path.exists():
            skipped_missing += 1
            print_progress(index, total_rows, f"missing {original_name}")
            continue

        desired_name = build_new_filename(
            row_number=row_number,
            sheet=sheet,
            headers=headers,
            source_name=args.name_source,
            base_name=args.base_name,
            sequence_number=sequence_number,
        )
        if desired_name.lower() == source_path.name.lower():
            target_path = source_path
            used_names.add(target_path.name.lower())
        else:
            target_path = unique_target_path(wav_dir, desired_name, used_names)
            source_path.rename(target_path)
            renamed_count += 1
        if args.name_source == "sequential":
            sequence_number += 1

        output_sheet.append(
            [
                target_path.name,
                sheet.cell(row=row_number, column=headers["old text"]).value,
                sheet.cell(row=row_number, column=headers["normalize text"]).value,
            ]
        )

        print_progress(index, total_rows, target_path.name)

    output_workbook.save(output_path)
    print()
    print(f"Done. Renamed {renamed_count} file(s).")
    print(f"Output Excel file created: {output_path}")
    if skipped_missing:
        print(f"Skipped {skipped_missing} row(s) with missing audio files or names.")


if __name__ == "__main__":
    main()
