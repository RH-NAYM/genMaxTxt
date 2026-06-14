from __future__ import annotations

import argparse
import json
import re
import os
import time
from pathlib import Path


NORMALIZATION_PROMPT = """You are normalizing Bangla ASR transcript text for text-to-speech.

Task:
- Convert only entities that need spoken Bangla normalization.
- Keep the original sentence meaning and natural conversational flow.
- Preserve normal Bangla words and ordinary punctuation.
- Return only the final normalized text. Do not explain anything.

Normalize these entity types:
1. English digits and Bangla digits, standalone or inside text.
2. Long numeric identifiers digit-by-digit: mobile/telephone numbers, OTP, PIN,
   verification code, NID, passport, driving license, bank/card/account numbers,
   tracking/order/invoice/reference/ticket/registration/serial/employee/student/
   customer/transaction/shipment IDs, and other long numeric sequences.
3. Phone numbers including Bangladesh mobile numbers, international numbers,
   country codes, landlines, hotlines. Example: +8801638830165 -> প্লাস আট আট
   শূন্য এক ছয় তিন আট আট তিন শূন্য এক ছয় পাঁচ.
4. Mathematical values: arithmetic expressions, percentages, decimals,
   fractions, measurements, quantities.
5. Emails: @ -> অ্যাট, . -> ডট, _ -> আন্ডারস্কোর, - -> হাইফেন,
   + -> প্লাস, / -> স্ল্যাশ. Spell English letters naturally in Bangla.
   Example: abc123@gmail.com -> এ বি সি এক দুই তিন অ্যাট জিমেইল ডট কম.
6. URLs/websites/domains/subdomains/query parameters. Example:
   https://example.com -> এইচ টি টি পি এস স্ল্যাশ স্ল্যাশ এক্সাম্পল ডট কম.
7. Dates in DD/MM/YYYY, MM/DD/YYYY, YYYY-MM-DD, written dates, Bangla dates,
   and relative dates. Use natural Bangla date style where clear.
8. Day names: Sunday through Saturday -> Bangla day names.
9. Month names: January through December -> natural Bangla pronunciation.
10. Years. Examples: 2025 -> দুই হাজার পঁচিশ সাল,
    1971 -> উনিশশো একাত্তর সাল.
11. Times: HH:MM, HH:MM:SS, AM/PM, 24-hour format.
12. Currency: BDT, USD, EUR, GBP, INR, SAR, AED and symbols.
13. Addresses: house, road, flat, postal code, area names.
14. Abbreviations: Dr., Mr., Mrs., Prof., Engr., Md., Ltd., Co.
15. Acronyms: AI, IT, ICT, USA, UK, NGO, API, GPU, CPU and similar forms.
16. Social media handles, hashtags, mentions.

Important style rules:
- For phone numbers and long IDs, read every digit separately.
- For ordinary small quantities, use natural Bangla number words.
- For ambiguous dates like 12/11/2025, keep a sensible spoken date without
  inventing extra context.
- Do not translate the whole sentence unless needed for fluent normalization.
- Do not add quotation marks around the answer.

Transcript:
{text}
"""

NORMALIZATION_PATTERNS = (
    re.compile(r"[0-9০-৯]"),
    re.compile(r"(?:https?://|www\.)", re.IGNORECASE),
    re.compile(r"\b[\w.%+-]+@[\w.-]+\.[A-Za-z]{2,}\b"),
    re.compile(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b"),
    re.compile(r"\b\d{4}-\d{1,2}-\d{1,2}\b"),
    re.compile(r"\b\d{1,2}:\d{2}(?::\d{2})?\s?(?:AM|PM|A\.M\.|P\.M\.)?\b", re.IGNORECASE),
    re.compile(r"[৳$€£₹]|\b(?:BDT|USD|EUR|GBP|INR|SAR|AED)\b", re.IGNORECASE),
    re.compile(r"(?<!\w)(?:\+?\d[\d\s().-০-৯]{6,}\d)(?!\w)"),
    re.compile(r"(?<!\w)(?:@[\w০-৯._-]+|#[\w০-৯._-]+)"),
    re.compile(r"\b(?:AI|IT|ICT|USA|UK|NGO|API|GPU|CPU)\b"),
    re.compile(r"\b(?:Dr|Mr|Mrs|Prof|Engr|Md|Ltd|Co)\.", re.IGNORECASE),
    re.compile(
        r"\b(?:Sunday|Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|"
        r"January|February|March|April|May|June|July|August|September|October|November|December)\b",
        re.IGNORECASE,
    ),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Normalize wav_text.xlsx transcripts with Gemini 2.5 Flash."
    )
    parser.add_argument(
        "--input",
        default="wav_text.xlsx",
        help="Input Excel file. Default: wav_text.xlsx",
    )
    parser.add_argument(
        "--file",
        dest="input",
        help="Input Excel file. Example: --file wav_text.xlsx",
    )
    parser.add_argument(
        "--output",
        default="wav_text_normalized_3_columns.xlsx",
        help="Output Excel file. Default: wav_text_normalized_3_columns.xlsx",
    )
    parser.add_argument(
        "--model",
        default="gemini-2.5-flash",
        help="Gemini model name. Default: gemini-2.5-flash",
    )
    parser.add_argument(
        "--api-key",
        help="Optional Gemini API key fallback. If omitted, service account Vertex AI is used.",
    )
    parser.add_argument(
        "--credentials",
        default="service_account.json",
        help="Google service account JSON file for Vertex AI. Default: service_account.json",
    )
    parser.add_argument(
        "--project",
        help="Google Cloud project ID. If omitted, read from service account JSON.",
    )
    parser.add_argument(
        "--location",
        default="us-central1",
        help="Vertex AI location for Gemini. Default: us-central1",
    )
    parser.add_argument(
        "--text-column",
        default="Text",
        help="Column containing transcript text. Default: Text",
    )
    parser.add_argument(
        "--id-column",
        default="Audio Id",
        help="Stable row ID column for resume. Default: Audio Id",
    )
    parser.add_argument(
        "--sheet",
        help="Input sheet name. If omitted, the active sheet is used.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Maximum number of input rows to process in this run.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="First input row number to process. Default: 2",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.0,
        help="Seconds to wait between Gemini calls. Default: 0",
    )
    parser.add_argument(
        "--save-every",
        type=int,
        default=1,
        help="Save output workbook after this many processed rows. Default: 1",
    )
    parser.add_argument(
        "--process-errors",
        action="store_true",
        help="Also send rows whose text starts with ERROR: or SKIPPED: to Gemini.",
    )
    parser.add_argument(
        "--max-retries",
        type=int,
        default=0,
        help="Maximum Gemini retries per row. Use 0 to retry forever. Default: 0",
    )
    parser.add_argument(
        "--retry-delay",
        type=float,
        default=15.0,
        help="Initial seconds to wait before retrying Gemini errors. Default: 15",
    )
    parser.add_argument(
        "--max-retry-delay",
        type=float,
        default=300.0,
        help="Maximum seconds to wait between Gemini retries. Default: 300",
    )
    return parser.parse_args()


def require_dependencies():
    try:
        from google import genai
        from google.genai import types
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        missing = exc.name or "required package"
        raise SystemExit(
            f"Missing dependency: {missing}\n"
            "Install dependencies with:\n"
            "  pip install google-genai openpyxl"
        ) from exc

    return genai, types, Workbook, load_workbook


def resolve_path(base_dir: Path, value: str) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return base_dir / path


def project_from_service_account(credentials_path: Path) -> str | None:
    try:
        with credentials_path.open("r", encoding="utf-8") as file:
            data = json.load(file)
    except (OSError, json.JSONDecodeError):
        return None
    project_id = data.get("project_id")
    return str(project_id) if project_id else None


def clean_model_text(value: str) -> str:
    text = value.strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        return text[1:-1].strip()
    return text


def is_retryable_gemini_error(exc: Exception) -> bool:
    message = str(exc).upper()
    retryable_markers = (
        "429",
        "RESOURCE_EXHAUSTED",
        "RATE_LIMIT",
        "RATE LIMIT",
        "QUOTA",
        "503",
        "UNAVAILABLE",
        "500",
        "INTERNAL",
        "504",
        "DEADLINE_EXCEEDED",
        "TIMEOUT",
        "CONNECTION",
    )
    return any(marker in message for marker in retryable_markers)


def needs_normalization(text: str) -> bool:
    return any(pattern.search(text) for pattern in NORMALIZATION_PATTERNS)


def normalize_text(client, types, model: str, text: str) -> str:
    prompt = NORMALIZATION_PROMPT.format(text=text)
    response = client.models.generate_content(
        model=model,
        contents=prompt,
        config=types.GenerateContentConfig(
            temperature=0.0,
            response_mime_type="text/plain",
        ),
    )
    normalized = getattr(response, "text", "") or ""
    normalized = clean_model_text(normalized)
    if not normalized:
        raise RuntimeError("Gemini returned an empty response")
    return normalized


def normalize_text_with_retries(
    client,
    types,
    model: str,
    text: str,
    max_retries: int,
    retry_delay: float,
    max_retry_delay: float,
    on_retry=None,
) -> str:
    attempt = 1
    delay = max(1.0, retry_delay)
    max_delay = max(delay, max_retry_delay)

    while True:
        try:
            return normalize_text(client, types, model, text)
        except Exception as exc:
            if not is_retryable_gemini_error(exc):
                raise
            if max_retries > 0 and attempt > max_retries:
                raise

            wait_seconds = min(delay, max_delay)
            if on_retry:
                on_retry(attempt, wait_seconds, exc)
            time.sleep(wait_seconds)
            attempt += 1
            delay = min(delay * 2, max_delay)


def header_map(sheet) -> dict[str, int]:
    headers = {}
    for column_number in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=column_number).value
        if value:
            headers[str(value).strip()] = column_number
    return headers


def output_headers(input_headers: list[str], text_column: str) -> list[str]:
    return ["audio file", "old text", "normalize text"]


def ensure_output_headers(sheet, expected_headers: list[str]) -> dict[str, int]:
    headers = header_map(sheet)
    for header in expected_headers:
        if header not in headers:
            column = sheet.max_column + 1
            sheet.cell(row=1, column=column).value = header
            headers[header] = column
    return headers


def get_or_create_output(output_path: Path, input_headers: list[str], args, Workbook, load_workbook):
    expected_headers = output_headers(input_headers, args.text_column)
    if output_path.exists():
        workbook = load_workbook(output_path)
        sheet = workbook.active
        existing_headers = [
            sheet.cell(row=1, column=column).value
            for column in range(1, sheet.max_column + 1)
        ]
        if existing_headers != expected_headers:
            raise SystemExit(
                f"Output file already exists with different columns: {output_path}\n"
                "Use --output with a new file name for the 3-column format."
            )
        return workbook, sheet

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "normalized_text"
    sheet.append(expected_headers)
    workbook.save(output_path)
    return workbook, sheet


def completed_keys(sheet, id_column: str) -> set[str]:
    headers = header_map(sheet)
    id_index = headers.get("audio file")
    old_text_index = headers.get("old text")
    text_index = headers.get("normalize text")
    if not id_index or not old_text_index or not text_index:
        return set()

    keys = set()
    for row_number in range(2, sheet.max_row + 1):
        row_id = sheet.cell(row=row_number, column=id_index).value
        old_text = str(sheet.cell(row=row_number, column=old_text_index).value or "").strip()
        normalized_text = sheet.cell(row=row_number, column=text_index).value
        normalized_text = str(normalized_text or "").strip()
        if row_id and not old_text:
            keys.add(str(row_id))
        elif row_id and normalized_text and not normalized_text.startswith("ERROR:"):
            keys.add(str(row_id))
    return keys


def output_rows_by_key(sheet) -> dict[str, int]:
    headers = header_map(sheet)
    id_index = headers.get("audio file")
    if not id_index:
        return {}

    rows = {}
    for row_number in range(2, sheet.max_row + 1):
        row_id = sheet.cell(row=row_number, column=id_index).value
        if row_id:
            rows[str(row_id)] = row_number
    return rows


def row_values(sheet, row_number: int, max_column: int) -> list:
    return [sheet.cell(row=row_number, column=column).value for column in range(1, max_column + 1)]


def set_cell_by_header(sheet, row_number: int, headers: dict[str, int], name: str, value) -> None:
    column = headers.get(name)
    if not column:
        raise KeyError(f"Output column not found: {name}")
    sheet.cell(row=row_number, column=column).value = value


def format_duration(seconds: float) -> str:
    seconds = max(0, int(seconds))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours:d}h {minutes:02d}m {seconds:02d}s"
    if minutes:
        return f"{minutes:d}m {seconds:02d}s"
    return f"{seconds:d}s"


def print_progress(
    done: int,
    total: int,
    started_at: float,
    current: str = "",
    status: str = "",
) -> None:
    width = 30
    filled = int(width * done / total) if total else width
    bar = "#" * filled + "-" * (width - filled)
    percent = (100 * done / total) if total else 100.0
    elapsed = time.time() - started_at
    if done and total and done < total:
        eta_seconds = (elapsed / done) * (total - done)
        eta = format_duration(eta_seconds)
    else:
        eta = "0s"
    message = (
        f"\rProgress: [{bar}] {done}/{total} ({percent:5.1f}%) "
        f"elapsed {format_duration(elapsed)} eta {eta}"
    )
    if status:
        message += f" {status}"
    if current:
        message += f" {current}"
    print(message[:180].ljust(180), end="", flush=True)


def main() -> None:
    args = parse_args()
    genai, types, Workbook, load_workbook = require_dependencies()

    base_dir = Path(__file__).resolve().parent
    input_path = resolve_path(base_dir, args.input)
    output_path = resolve_path(base_dir, args.output)
    credentials_path = resolve_path(base_dir, args.credentials)

    if not input_path.exists():
        raise SystemExit(f"Input Excel file not found: {input_path}")
    if not args.api_key and not credentials_path.exists():
        raise SystemExit(f"Google credentials file not found: {credentials_path}")
    if args.start_row < 2:
        raise SystemExit("--start-row must be 2 or greater")
    if args.save_every < 1:
        raise SystemExit("--save-every must be 1 or greater")
    if args.max_retries < 0:
        raise SystemExit("--max-retries must be 0 or greater")
    if args.retry_delay <= 0:
        raise SystemExit("--retry-delay must be greater than 0")
    if args.max_retry_delay <= 0:
        raise SystemExit("--max-retry-delay must be greater than 0")

    input_workbook = load_workbook(input_path, read_only=True)
    if args.sheet:
        if args.sheet not in input_workbook.sheetnames:
            raise SystemExit(f"Sheet not found: {args.sheet}")
        input_sheet = input_workbook[args.sheet]
    else:
        input_sheet = input_workbook.active

    input_header_lookup = header_map(input_sheet)
    text_index = input_header_lookup.get(args.text_column)
    if not text_index:
        raise SystemExit(f"Text column not found: {args.text_column}")

    id_index = input_header_lookup.get(args.id_column)
    input_headers = [
        str(input_sheet.cell(row=1, column=column).value or f"Column {column}")
        for column in range(1, input_sheet.max_column + 1)
    ]

    output_workbook, output_sheet = get_or_create_output(
        output_path, input_headers, args, Workbook, load_workbook
    )
    output_header_lookup = header_map(output_sheet)
    done_keys = completed_keys(output_sheet, args.id_column)
    output_row_lookup = output_rows_by_key(output_sheet)

    if args.api_key:
        client = genai.Client(api_key=args.api_key)
    else:
        project = args.project or project_from_service_account(credentials_path)
        if not project:
            raise SystemExit(
                "Google Cloud project ID not found. Pass --project or use a service "
                "account JSON with project_id."
            )
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = str(credentials_path)
        client = genai.Client(vertexai=True, project=project, location=args.location)

    candidates = []
    for row_number in range(args.start_row, input_sheet.max_row + 1):
        row_id = input_sheet.cell(row=row_number, column=id_index).value if id_index else row_number
        key = str(row_id)
        if key in done_keys:
            continue
        candidates.append(row_number)
        if args.limit and len(candidates) >= args.limit:
            break

    started_at = time.time()
    print_progress(0, len(candidates), started_at)
    processed_since_save = 0

    for done, row_number in enumerate(candidates, start=1):
        values = row_values(input_sheet, row_number, input_sheet.max_column)
        original_text = str(input_sheet.cell(row=row_number, column=text_index).value or "").strip()
        row_id = values[id_index - 1] if id_index else row_number
        current_name = str(row_id)

        output_row = output_row_lookup.get(current_name)
        if output_row:
            set_cell_by_header(output_sheet, output_row, output_header_lookup, "audio file", row_id)
            set_cell_by_header(output_sheet, output_row, output_header_lookup, "old text", original_text)
            set_cell_by_header(output_sheet, output_row, output_header_lookup, "normalize text", "")
        else:
            output_sheet.append([row_id, original_text, ""])
            output_row = output_sheet.max_row
            output_row_lookup[current_name] = output_row

        if not original_text:
            set_cell_by_header(output_sheet, output_row, output_header_lookup, "normalize text", "")
            row_status = "EMPTY"
        elif not needs_normalization(original_text):
            set_cell_by_header(output_sheet, output_row, output_header_lookup, "normalize text", original_text)
            row_status = "COPY"
        elif (
            not args.process_errors
            and (original_text.startswith("ERROR:") or original_text.startswith("SKIPPED:"))
        ):
            set_cell_by_header(output_sheet, output_row, output_header_lookup, "normalize text", original_text)
            row_status = "SKIPPED"
        else:
            try:
                def show_retry(attempt: int, wait_seconds: float, exc: Exception) -> None:
                    retry_status = f"RETRY {attempt} wait {format_duration(wait_seconds)}"
                    print_progress(done - 1, len(candidates), started_at, current_name, retry_status)

                normalized = normalize_text_with_retries(
                    client,
                    types,
                    args.model,
                    original_text,
                    args.max_retries,
                    args.retry_delay,
                    args.max_retry_delay,
                    show_retry,
                )
                set_cell_by_header(output_sheet, output_row, output_header_lookup, "normalize text", normalized)
                row_status = "DONE"
            except Exception as exc:
                set_cell_by_header(
                    output_sheet,
                    output_row,
                    output_header_lookup,
                    "normalize text",
                    f"ERROR: {exc}",
                )
                row_status = "ERROR"

        processed_since_save += 1
        if processed_since_save >= args.save_every:
            output_workbook.save(output_path)
            processed_since_save = 0

        print_progress(done, len(candidates), started_at, current_name, row_status)
        if args.sleep > 0 and done < len(candidates):
            time.sleep(args.sleep)

    if processed_since_save:
        output_workbook.save(output_path)

    print()
    print(f"Done. Normalized Excel file created: {output_path}")


if __name__ == "__main__":
    main()
