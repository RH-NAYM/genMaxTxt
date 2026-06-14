from __future__ import annotations

import argparse
import io
import os
import struct
import wave
from pathlib import Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an Excel file with wav file names and Google STT text."
    )
    parser.add_argument(
        "--wav-dir",
        default="wavs",
        help="Folder containing .wav files. Default: wavs",
    )
    parser.add_argument(
        "--dir",
        dest="wav_dir",
        help="Folder containing .wav files. Example: --dir wavs",
    )
    parser.add_argument(
        "--file",
        help="Single .wav file to transcribe. Example: --file wavs\\294.wav",
    )
    parser.add_argument(
        "--output",
        default="wav_text.xlsx",
        help="Output Excel file name. Default: wav_text.xlsx",
    )
    parser.add_argument(
        "--credentials",
        default="service_account.json",
        help="Google service account JSON file. Default: service_account.json",
    )
    parser.add_argument(
        "--language",
        default="bn-BD",
        help="Speech language code. Default: bn-BD",
    )
    return parser.parse_args()


def require_dependencies():
    try:
        from google.cloud import speech
        from openpyxl import load_workbook
        from openpyxl import Workbook
    except ImportError as exc:
        missing = exc.name or "required package"
        raise SystemExit(
            f"Missing dependency: {missing}\n"
            "Install dependencies with:\n"
            "  pip install google-cloud-speech openpyxl"
        ) from exc

    return speech, Workbook, load_workbook


def read_wav_chunks(wav_path: Path) -> tuple[dict[str, int], bytes]:
    data = wav_path.read_bytes()
    if data[:4] != b"RIFF" or data[8:12] != b"WAVE":
        raise ValueError("Not a valid WAV file")

    info = {}
    audio_data = b""
    position = 12
    while position + 8 <= len(data):
        chunk_id = data[position : position + 4]
        chunk_size = struct.unpack_from("<I", data, position + 4)[0]
        chunk_start = position + 8
        chunk_end = chunk_start + chunk_size
        chunk = data[chunk_start:chunk_end]

        if chunk_id == b"fmt ":
            (
                info["audio_format"],
                info["channels"],
                info["sample_rate"],
                _byte_rate,
                _block_align,
                info["bits_per_sample"],
            ) = struct.unpack_from("<HHIIHH", chunk)
        elif chunk_id == b"data":
            audio_data = chunk

        position = chunk_end + (chunk_size % 2)

    if not info or not audio_data:
        raise ValueError("WAV file is missing fmt or data chunk")

    return info, audio_data


def float_wav_to_linear16_content(info: dict[str, int], audio_data: bytes) -> bytes:
    if info["bits_per_sample"] != 32:
        raise ValueError(
            f"Unsupported float WAV bit depth: {info['bits_per_sample']}. "
            "Only 32-bit float WAV conversion is supported."
        )

    samples = struct.iter_unpack("<f", audio_data)
    pcm = bytearray()
    for (sample,) in samples:
        sample = max(-1.0, min(1.0, sample))
        value = int(sample * 32767) if sample >= 0 else int(sample * 32768)
        pcm.extend(struct.pack("<h", value))

    buffer = io.BytesIO()
    with wave.open(buffer, "wb") as wav_file:
        wav_file.setnchannels(info["channels"])
        wav_file.setsampwidth(2)
        wav_file.setframerate(info["sample_rate"])
        wav_file.writeframes(bytes(pcm))
    return buffer.getvalue()


def prepare_wav_for_google(wav_path: Path) -> tuple[bytes, int, int]:
    info, audio_data = read_wav_chunks(wav_path)
    audio_format = info["audio_format"]
    sample_rate_hertz = info["sample_rate"]
    audio_channel_count = info["channels"]

    if audio_format == 1 and info["bits_per_sample"] == 16:
        return wav_path.read_bytes(), sample_rate_hertz, audio_channel_count

    if audio_format == 3:
        content = float_wav_to_linear16_content(info, audio_data)
        return content, sample_rate_hertz, audio_channel_count

    raise ValueError(
        f"Unsupported WAV format: audio_format={audio_format}, "
        f"bits_per_sample={info['bits_per_sample']}"
    )


def transcribe_wav(client, speech, wav_path: Path, language_code: str) -> str:
    content, sample_rate_hertz, audio_channel_count = prepare_wav_for_google(wav_path)

    audio = speech.RecognitionAudio(content=content)
    config = speech.RecognitionConfig(
        encoding=speech.RecognitionConfig.AudioEncoding.LINEAR16,
        sample_rate_hertz=sample_rate_hertz,
        audio_channel_count=audio_channel_count,
        language_code=language_code,
        enable_automatic_punctuation=True,
    )

    response = client.recognize(config=config, audio=audio)
    return " ".join(
        result.alternatives[0].transcript.strip()
        for result in response.results
        if result.alternatives
    )


def print_progress(done: int, total: int, current_name: str = "") -> None:
    width = 30
    filled = int(width * done / total) if total else width
    bar = "#" * filled + "-" * (width - filled)
    percent = int(100 * done / total) if total else 100
    message = f"\rProgress: [{bar}] {done}/{total} ({percent}%)"
    if current_name:
        message += f" {current_name}"
    print(message[:120].ljust(120), end="", flush=True)


def get_or_create_workbook(output_path: Path, Workbook, load_workbook):
    if output_path.exists():
        workbook = load_workbook(output_path)
        sheet = workbook.active
        if sheet.max_row == 0:
            sheet.append(["Audio Id", "Text"])
        return workbook, sheet

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "wav_text"
    sheet.append(["Audio Id", "Text"])
    workbook.save(output_path)
    return workbook, sheet


def get_existing_rows(sheet) -> dict[str, tuple[int, str]]:
    rows = {}
    for row_number in range(2, sheet.max_row + 1):
        wav_name = sheet.cell(row=row_number, column=1).value
        wav_text = sheet.cell(row=row_number, column=2).value
        if wav_name:
            rows[str(wav_name)] = (row_number, str(wav_text or ""))
    return rows


def main() -> None:
    args = parse_args()
    speech, Workbook, load_workbook = require_dependencies()

    base_dir = Path(__file__).resolve().parent
    wav_dir = Path(args.wav_dir)
    wav_file = Path(args.file) if args.file else None
    output_path = Path(args.output)
    credentials_path = Path(args.credentials)

    if not wav_dir.is_absolute():
        wav_dir = base_dir / wav_dir
    if wav_file and not wav_file.is_absolute():
        wav_file = base_dir / wav_file
    if not output_path.is_absolute():
        output_path = base_dir / output_path
    if not credentials_path.is_absolute():
        credentials_path = base_dir / credentials_path

    if wav_file:
        if not wav_file.exists():
            raise SystemExit(f"WAV file not found: {wav_file}")
        if wav_file.suffix.lower() != ".wav":
            raise SystemExit(f"Not a .wav file: {wav_file}")
        wav_files = [wav_file]
    else:
        if not wav_dir.exists():
            raise SystemExit(f"WAV folder not found: {wav_dir}")
        wav_files = sorted(wav_dir.rglob("*.wav"))

    if not credentials_path.exists():
        raise SystemExit(f"Google credentials file not found: {credentials_path}")

    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = str(credentials_path)

    if not wav_files:
        raise SystemExit(f"No .wav files found in: {wav_dir}")

    client = speech.SpeechClient()
    workbook, sheet = get_or_create_workbook(output_path, Workbook, load_workbook)
    existing_rows = get_existing_rows(sheet)
    completed_count = sum(
        1
        for wav_path in wav_files
        if wav_path.name in existing_rows and existing_rows[wav_path.name][1]
    )

    print_progress(completed_count, len(wav_files))

    for wav_path in wav_files:
        existing = existing_rows.get(wav_path.name)
        if existing and existing[1]:
            print_progress(completed_count, len(wav_files), f"skipped {wav_path.name}")
            continue

        print()
        print(f"Transcribing: {wav_path.name}")
        try:
            text = transcribe_wav(client, speech, wav_path, args.language)
        except Exception as exc:
            text = f"ERROR: {exc}"
        print(f"File==>>>> {wav_path.name}, Text==>>>> {text}")

        if existing:
            row_number = existing[0]
            sheet.cell(row=row_number, column=2).value = text
        else:
            row_number = sheet.max_row + 1
            sheet.append([wav_path.name, text])

        workbook.save(output_path)
        existing_rows[wav_path.name] = (row_number, text)
        completed_count += 1
        print_progress(completed_count, len(wav_files), f"saved {wav_path.name}")

    print()
    print(f"Done. Excel file created: {output_path}")


if __name__ == "__main__":
    main()
